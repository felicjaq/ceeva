from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.certificates import format_cert_info, extract_cert_info
from datetime import datetime

def calculate_days_until_expiry(expiry_date):
    expiry = datetime.strptime(expiry_date, '%Y-%m-%d %H:%M:%S')
    remaining_days = (expiry - datetime.now()).days
    return remaining_days

def get_fill_color_for_expiry(remaining_days):
    if remaining_days < 0:
        return 'FFC7CE'
    elif remaining_days < 7:
        return 'FFCC99'
    elif remaining_days < 30:
        return 'FFEB9C'
    else:
        return 'C6EFCE'

def generate_excel_report(domains_certs, filename, language):
    wb = Workbook()
    ws = wb.active
    ws.title = language.get('sheet_title', 'Sheet')

    headers = [
        language['domain_ip'], language['port'], language['issued_date'], language['expiry_date'],
        language['days_until_expiry'], language['subject'], language['issuer'], language['serial_number']
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = header_border

    rows = []
    for domain, certs in domains_certs.items():
        for port, cert in certs:
            cert_info = format_cert_info(extract_cert_info(cert), language)
            remaining_days = calculate_days_until_expiry(cert_info[language['expiry_date']])
            row = [
                domain,
                port,
                cert_info[language['issued_date']],
                cert_info[language['expiry_date']],
                remaining_days,
                cert_info[language['subject']],
                cert_info[language['issuer']],
                str(cert_info[language['serial_number']])
            ]
            rows.append((remaining_days, row))

    rows.sort(key=lambda x: x[0])

    cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for _, row_data in rows:
        ws.append(row_data)
        fill_color = get_fill_color_for_expiry(row_data[4])
        expiry_cell = ws.cell(row=ws.max_row, column=5)
        expiry_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

    column_widths = [25, 10, 25, 25, 20, 30, 30, 25]
    for i, width in enumerate(column_widths):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = width

    wb.save(filename)
